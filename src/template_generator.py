"""
template_generator.py
─────────────────────
Creates the Leave_Tracker_Template.xlsx with two sheets:
  • Employee Master
  • Leave Tracker

Uses only dummy placeholder data — no real names or emails.

Author : Debojit Dhali
License: MIT
"""

from __future__ import annotations

import os
from datetime import date
from typing import Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Style helpers ─────────────────────────────────────────────────────────────

def _thin_border(color: str = "CCCCCC") -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_cell(cell, bg: str, fg: str = "FFFFFF") -> None:
    cell.font      = Font(bold=True, color=fg, size=11, name="Segoe UI")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _thin_border("BBBBBB")


def _data_cell(cell, bg: str = "FFFFFF") -> None:
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.border    = _thin_border("DDDDDD")
    cell.font      = Font(name="Segoe UI", size=10)


# ── Public API ────────────────────────────────────────────────────────────────

def generate(output_dir: str) -> Tuple[bool, str, str]:
    """
    Generate the template Excel file.
    Returns (success, output_path, error_message).
    """
    try:
        wb  = openpyxl.Workbook()
        _build_employee_master(wb)
        _build_leave_tracker(wb)

        out_path = os.path.join(output_dir, "Leave_Tracker_Template.xlsx")
        wb.save(out_path)
        return True, out_path, ""

    except Exception as e:
        return False, "", str(e)


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_employee_master(wb: openpyxl.Workbook) -> None:
    ws = wb.active
    ws.title = "Employee Master"

    headers = ["Name", "Email", "Team", "Manager Name", "Manager Email", "CC1", "CC2", "CC3"]
    widths  = [22,     30,      16,     22,              30,              26,    26,    26   ]

    # ── Header row ────────────────────────────────────────────────────────────
    ws.append(headers)
    for i, (cell, w) in enumerate(zip(ws[1], widths), start=1):
        _header_cell(cell, bg="1e3a5f")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 26

    # ── Sample rows (dummy data only) ─────────────────────────────────────────
    rows = [
        ["Firstname1 Lastname1", "firstname1.lastname1@company.com", "Dev Team",
         "Lead1 Name",    "lead1@company.com",    "scrummaster@company.com", "pm@company.com",   ""],
        ["Firstname2 Lastname2", "firstname2.lastname2@company.com", "Testing Team",
         "Lead2 Name",    "lead2@company.com",    "scrummaster@company.com", "",                 ""],
        ["Firstname3 Lastname3", "firstname3.lastname3@company.com", "DB Team",
         "Manager1 Name", "manager1@company.com", "scrummaster@company.com", "pm@company.com",   "tl@company.com"],
        ["Firstname4 Lastname4", "firstname4.lastname4@company.com", "Dev Team",
         "Lead1 Name",    "lead1@company.com",    "scrummaster@company.com", "",                 ""],
        ["Firstname5 Lastname5", "firstname5.lastname5@company.com", "Testing Team",
         "Lead2 Name",    "lead2@company.com",    "scrummaster@company.com", "pm@company.com",   ""],
    ]

    alt = ["F0F4FF", "FFFFFF"]
    for r_idx, row_data in enumerate(rows, start=2):
        ws.append(row_data)
        bg = alt[(r_idx - 2) % 2]
        for cell in ws[r_idx]:
            _data_cell(cell, bg=bg)
        ws.row_dimensions[r_idx].height = 18

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def _build_leave_tracker(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Leave Tracker")

    headers = ["Employee Name",  "Leave Type", "Start Date", "End Date", "Reason"]
    widths  = [26,                16,            16,           16,         34     ]

    ws.append(headers)
    for i, (cell, w) in enumerate(zip(ws[1], widths), start=1):
        _header_cell(cell, bg="0c4a6e")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 26

    today     = date.today().strftime("%d-%m-%Y")
    leave_bg  = {"Unplanned": "FFF3CD", "Planned": "D1F2EB"}

    rows = [
        ["Firstname1 Lastname1", "Unplanned", today, today, "Sick leave"],
        ["Firstname3 Lastname3", "Planned",   today, today, "Personal work"],
    ]

    for r_idx, row_data in enumerate(rows, start=2):
        ws.append(row_data)
        bg = leave_bg.get(str(row_data[1]), "FFFFFF")
        for cell in ws[r_idx]:
            _data_cell(cell, bg=bg)
        ws.row_dimensions[r_idx].height = 18

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
